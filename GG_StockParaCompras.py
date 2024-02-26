import pyodbc
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

# Establecer la cadena de conexión al servidor SQL Server
conn_str = 'DRIVER={SQL Server};SERVER=192.168.2.20;DATABASE=GaleriaGrafica;UID=GaleriaGrafica;PWD=Galeria123'

# Conectarse al servidor de base de datos
conn = pyodbc.connect(conn_str)

# Crear un cursor para ejecutar consultas
cursor = conn.cursor()

# Ejecutar la consulta SQL
query = """
SELECT *
FROM (
    SELECT LN.OP AS "Necesidad Generada Por",
           P.OP,
           P.OPA,
           LN.Articulo AS Codigo_Material,
           LN.Descripcion AS Material,
           LAA.Stk_Actual AS STK_Actual,
           ISNULL(CD.STK_Disponivel, 0) AS STK_Disponivel,
           CD.STK_Reservado,
           CD.STK_Encomendado,
           LN.Cantidad AS CantidadSolicitidad,
           LN.[Cantidad Abierto] AS CantidadAbierta,
           P.Cliente,
           P.Referencia_Articulo,
           P.Descripcion_Articulo,
           P.Estado_OP,
           P.Fecha_Solicitada_Cliente,
           ROW_NUMBER() OVER (PARTITION BY LN.OP, LN.Articulo ORDER BY LN.OP, LN.Articulo) AS RowNum
    FROM (
        SELECT CONCAT([OP Serie], '/', [OP Codigo]) AS OP, Articulo, Descripcion, Cantidad, UM, [Cantidad Abierto]
        FROM [00GG_Lista_Necesidades]
    ) AS LN
    JOIN [00GG_PEDIDOS] AS P ON LN.OP = CASE WHEN LEN(P.OPA) > 0 THEN P.OPA ELSE P.OP END
    JOIN [00GG_LISTADO_ARTICULOS] AS LAA ON LAA.Codigo = LN.Articulo
    LEFT JOIN [00GG_CONSULTA_DISPONIBLE] AS CD ON CD.Codigo_Articulo = LN.Articulo AND LN.[Cantidad Abierto] > 0
    WHERE P.Estado_OP IN ('(C)Registrada', '(O)Aprobada Producción', '(D)Pendiente de Aprobación', '(P)Producción')
      AND (LN.Articulo LIKE '10%' OR LN.Articulo LIKE '15%' OR LN.Articulo LIKE '20%' OR LN.Articulo LIKE '25%' OR LN.Articulo LIKE '28%' OR LN.Articulo LIKE '69%' OR LN.Articulo LIKE '77%')
      AND LN.[Cantidad Abierto] > 0
) AS Subquery
WHERE RowNum = 1
ORDER BY Material, Estado_OP DESC, CONVERT(DATE, Fecha_Solicitada_Cliente, 105) ASC;

"""

cursor.execute(query)

# Obtener los resultados de la consulta
results = cursor.fetchall()

# Crear un nuevo archivo de Excel
wb = openpyxl.Workbook()
ws = wb.active

# Establecer estilos para colorear las celdas
green_fill = PatternFill(start_color='66D56D', end_color='66D56D', fill_type='solid')
red_fill = PatternFill(start_color='F1533D', end_color='F1533D', fill_type='solid')
orange_fill = PatternFill(start_color='FCD884', end_color='FCD884', fill_type='solid')

# Escribir los encabezados de las columnas en la hoja de Excel
headers = [column[0] for column in cursor.description]
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws[f'{col_letter}1'] = header

# Calcular la disponibilidad y aplicar formato de color por material

current_material = None
stk_actual = None
stk_encomendado = None
cantidad_abierta_total = 0

for row_num, row in enumerate(results, 2):
    if row[4] != current_material:
        # Cambio de material, actualizar las variables
        current_material = row[4]
        stk_actual = row[5]
        stk_encomendado = row[8]
        cantidad_abierta_total = 0

    cantidad_abierta = row[10]
    cantidad_abierta_total += cantidad_abierta
    estado_op = row[14]

    if cantidad_abierta_total <= stk_actual:
        # Suficiente stock disponible (STK_Actual) para cubrir la cantidad solicitada
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.fill = green_fill
    elif cantidad_abierta_total <= (stk_actual + stk_encomendado):
        # Se necesita el stock encomendado (STK_Encomendado) para cubrir la cantidad solicitada
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.fill = orange_fill
    else:
        # No hay suficiente stock disponible ni encomendado para cubrir la cantidad solicitada
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.fill = red_fill

    # Escribir los datos en la hoja de Excel
    for col_num, value in enumerate(row, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}{row_num}'] = value

# Guardar el archivo de Excel
wb.save('resultados.xlsx')

# Cerrar la conexión y el cursor
cursor.close()
conn.close()
